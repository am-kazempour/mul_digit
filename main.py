from openpyxl import load_workbook,Workbook

file_names = [
    'Month1.xlsx','Month2.xlsx','Month3.xlsx',
    'Month4.xlsx','Month5.xlsx','Month6.xlsx',
    'Month7.xlsx','Month8.xlsx','Month9.xlsx',
    'Month10.xlsx','Month11.xlsx','Month12.xlsx',
]

# read files
workbooks = []
worksheets = []
for i,file in enumerate(file_names):
    workbooks.append(load_workbook('datasets/'+file))
    worksheets.append(workbooks[i].active)
    workbooks[i].close()

# fetch data
workers_name =[]
datas = []
for month in range(len(file_names)):
    Month_data = []
    for row in worksheets[month].iter_rows(values_only=True):
        Month_data.append(list(row))
        if len(Month_data) != 1:
            workers_name.insert(0,row[0])
    datas.append(Month_data[1:])

workers_name = sorted(list(set(workers_name)))

output = {name:[0 for _ in range(24)]for name in workers_name}

for i,data in enumerate(datas):
    for worker in data:
        output[worker[0]][2*i] = worker[1]
        output[worker[0]][2*i + 1] = worker[2]

workbook = Workbook()
worksheet = workbook.active

worksheet.append(['Name'] + [f'month {i+1} ({prefix})' for i in range(12) for prefix in ['day', 'salary']])

for name in output.keys():
    if output[name][-1] != 0:
        worksheet.append([name] + output[name])

for name in output.keys():
    if output[name][-1] == 0:
        worksheet.append([name] + output[name])

workbook.save('output.xlsx')
workbook.close()